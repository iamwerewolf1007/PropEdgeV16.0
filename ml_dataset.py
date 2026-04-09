"""
PropEdge V16.0 — ml_dataset.py
================================
Writes and appends the PropEdge ML training dataset to Excel.

Output: data/propedge_ml_dataset.xlsx

Two sheets:
  Plays  — one row per prop, all features + outcome + reasoning
  Schema — column definitions and group descriptions

Column groups (colour-coded header):
  A. IDENTITY     — player, date, team, season context
  B. MARKET       — line, direction, odds, books, line spread
  C. FORM         — L3/L5/L10/L20/L30, std, hit rates, trends
  D. PHYSICAL     — minutes, shooting efficiency, usage rate
  E. CONTEXT      — home/away, rest, DVP, pace, season progress
  F. H2H          — head-to-head history vs this specific opponent
  G. MODEL        — clf probabilities, trust scores, quantile bands
  H. PREDICTION   — predPts, tier, confidence, signal flags
  I. OUTCOME      — actual_pts, result, delta, box score stats
  J. ML_SIGNALS   — derived fields specifically for model training
  K. REASONING    — pre-match and post-match narrative text

Usage:
  from ml_dataset import write_ml_dataset, append_ml_dataset
  write_ml_dataset(plays)          # full rebuild (generate)
  append_ml_dataset(plays, date)   # daily append (batch0_grade)
"""

from __future__ import annotations

import math
from datetime import date as date_type
from pathlib import Path
from typing import Any

import pandas as pd

ROOT = Path(__file__).resolve().parent
FILE_ML = ROOT / "data" / "propedge_ml_dataset.xlsx"

# ── Colour palette for header groups ──────────────────────────────────────────
_C = {
    "A": "BDD7EE",  # Identity     — steel blue
    "B": "FFF2CC",  # Market       — warm yellow
    "C": "E2EFDA",  # Form         — soft green
    "D": "FCE4D6",  # Physical     — peach
    "E": "EAD1DC",  # Context      — lavender
    "F": "D9E1F2",  # H2H          — periwinkle
    "G": "D9D9D9",  # Model        — light grey
    "H": "FFD966",  # Prediction   — gold
    "I_win":  "C6EFCE",  # Outcome WIN  — green
    "I_loss": "FFC7CE",  # Outcome LOSS — red
    "I":      "F2F2F2",  # Outcome base — near white
    "J": "EDEDED",  # ML Signals   — grey
    "K": "FFFFFF",  # Reasoning    — white
}

# ── Column schema ─────────────────────────────────────────────────────────────
# Each entry: (field_key, header_label, group, width, fmt)
# fmt: None=general, "pct"=percentage, "int"=integer, "dec2"=2dp, "dec4"=4dp

COLUMNS: list[tuple[str, str, str, int, str | None]] = [
    # ── A. IDENTITY ──────────────────────────────────────────────────────────
    ("date",             "Date",              "A", 12, None),
    ("player",           "Player",            "A", 22, None),
    ("position",         "Position",          "A",  6, None),
    ("team",             "Team",              "A",  6, None),
    ("opponent",         "Opponent",          "A",  6, None),
    ("is_home",          "Home",              "A",  6, None),
    ("match",            "Match",             "A", 24, None),
    ("game_time",        "Game Time (ET)",    "A", 14, None),
    ("season",           "Season",            "A", 10, None),
    ("source",           "Source",            "A",  8, None),

    # ── B. MARKET ─────────────────────────────────────────────────────────────
    ("line",             "Line",              "B",  7, "dec1"),
    ("direction",        "Direction",         "B",  9, None),
    ("over_odds",        "Over Odds",         "B",  9, "int"),
    ("under_odds",       "Under Odds",        "B",  9, "int"),
    ("books",            "Books",             "B",  6, "int"),
    ("min_line",         "Min Line",          "B",  8, "dec1"),
    ("max_line",         "Max Line",          "B",  8, "dec1"),
    ("line_spread",      "Line Spread",       "B",  9, "dec1"),

    # ── C. FORM ───────────────────────────────────────────────────────────────
    ("l3",               "L3",                "C",  7, "dec1"),
    ("l5",               "L5",                "C",  7, "dec1"),
    ("l10",              "L10",               "C",  7, "dec1"),
    ("l20",              "L20",               "C",  7, "dec1"),
    ("l30",              "L30",               "C",  7, "dec1"),
    ("std10",            "Std10",             "C",  7, "dec1"),
    ("hr10",             "HR10",              "C",  7, "pct"),
    ("hr30",             "HR30",              "C",  7, "pct"),
    ("volume",           "Volume",            "C",  8, "dec1"),
    ("momentum",         "Momentum",          "C",  9, "dec1"),
    ("reversion",        "Reversion",         "C",  9, "dec1"),
    ("acceleration",     "Accel",             "C",  7, "dec1"),
    ("level_ewm",        "EWM L10",           "C",  8, "dec1"),
    ("mean_reversion_risk", "MRevRisk",       "C",  9, "dec2"),
    ("extreme_hot",      "Ext Hot",           "C",  7, None),
    ("extreme_cold",     "Ext Cold",          "C",  7, None),

    # ── D. PHYSICAL ───────────────────────────────────────────────────────────
    ("min_l10",          "Min L10",           "D",  8, "dec1"),
    ("min_l30",          "Min L30",           "D",  8, "dec1"),
    ("min_cv",           "Min CV",            "D",  7, "dec2"),
    ("pts_per_min",      "Pts/Min",           "D",  7, "dec2"),
    ("fga_l10",          "FGA L10",           "D",  8, "dec1"),
    ("fg3a_l10",         "3PA L10",           "D",  8, "dec1"),
    ("l10_fg_pct",       "FG% L10",           "D",  8, "pct"),
    ("ft_rate",          "FT Rate",           "D",  8, "pct"),
    ("fta_l10",          "FTA L10",           "D",  8, "dec1"),
    ("usage_l10",        "Usage L10",         "D",  9, "pct"),
    ("usage_l30",        "Usage L30",         "D",  9, "pct"),

    # ── E. CONTEXT ────────────────────────────────────────────────────────────
    ("homeAvgPts",       "Home Avg",          "E",  9, "dec1"),
    ("awayAvgPts",       "Away Avg",          "E",  9, "dec1"),
    ("home_away_split",  "H/A Split",         "E",  9, "dec1"),
    ("rest_days",        "Rest Days",         "E",  9, "int"),
    ("is_b2b",           "B2B",               "E",  5, None),
    ("is_long_rest",     "Long Rest",         "E",  9, None),
    ("defP_dynamic",     "DVP Rank",          "E",  9, "int"),
    ("pace_rank",        "Pace Rank",         "E",  9, "int"),
    ("season_progress",  "Season%",           "E",  8, "pct"),
    ("early_season_weight", "Early Wt",       "E",  8, "dec2"),

    # ── F. H2H ────────────────────────────────────────────────────────────────
    ("h2h_games",        "H2H Games",         "F", 10, "int"),
    ("h2h_avg",          "H2H Avg Pts",       "F", 11, "dec1"),
    ("h2hTsDev",         "H2H TS Dev",        "F", 10, "dec2"),
    ("h2hFgaDev",        "H2H FGA Dev",       "F", 10, "dec2"),
    ("h2hConfidence",    "H2H Conf",          "F",  9, "dec2"),
    ("h2h_trend",        "H2H Trend",         "F", 10, "dec2"),

    # ── G. MODEL ──────────────────────────────────────────────────────────────
    ("v12_clf_prob",     "V12 Clf Prob",      "G", 11, "dec4"),
    ("v14_clf_prob",     "V14 Clf Prob",      "G", 11, "dec4"),
    ("v12_clf_conv",     "V12 Conviction",    "G", 12, "dec3"),
    ("real_gap_v92",     "Gap V92",           "G",  9, "dec2"),
    ("real_gap_v12",     "Gap V12",           "G",  9, "dec2"),
    ("real_gap_mean",    "Gap Mean",          "G",  9, "dec2"),
    ("q25_v12",          "Q25 V12",           "G",  9, "dec2"),
    ("q75_v12",          "Q75 V12",           "G",  9, "dec2"),
    ("q_confidence",     "Q Conf",            "G",  8, "dec3"),
    ("trust_v12",        "Trust V12",         "G", 10, "dec3"),
    ("trust_v14",        "Trust V14",         "G", 10, "dec3"),
    ("trust_mean",       "Trust Mean",        "G", 10, "dec3"),
    ("all_clf_agree",    "All Agree",         "G", 10, None),
    ("reg_consensus",    "Reg Consensus",     "G", 12, None),
    ("v12_extreme",      "V12 Extreme",       "G", 11, None),
    ("flags",            "Signal Flags",      "G", 11, "int"),

    # ── H. PREDICTION ─────────────────────────────────────────────────────────
    ("predPts",          "Pred Pts",          "H",  9, "dec1"),
    ("predGap",          "Pred Gap",          "H",  9, "dec2"),
    ("elite_prob",       "Elite Prob",        "H", 10, "dec4"),
    ("elite_tier",       "Elite Tier",        "H", 10, None),
    ("elite_stake",      "Stake (u)",         "H",  9, "dec1"),
    ("conf",             "Conf",              "H",  8, "dec4"),

    # ── I. OUTCOME ────────────────────────────────────────────────────────────
    ("result",           "Result",            "I",  8, None),
    ("actualPts",        "Actual Pts",        "I", 10, "dec1"),
    ("delta",            "Delta",             "I",  7, "dec1"),
    ("actual_min",       "Actual Min",        "I", 10, "dec1"),
    ("actual_fga",       "Actual FGA",        "I", 10, "dec1"),
    ("actual_fgm",       "Actual FGM",        "I", 10, "dec1"),
    ("actual_fg_pct",    "Actual FG%",        "I", 10, "pct"),
    ("lossType",         "Loss Type",         "I", 18, None),

    # ── J. ML SIGNALS (derived) ───────────────────────────────────────────────
    ("ml_correct",       "ML Correct",        "J", 10, None),
    ("ml_pred_error",    "Pred Error",        "J", 10, "dec1"),
    ("ml_dir_correct",   "Dir Correct",       "J", 11, None),
    ("ml_close_call",    "Close Call",        "J", 10, None),
    ("ml_within_q",      "In Q Band",         "J", 10, None),
    ("ml_consensus",     "Consensus Scr",     "J", 13, "int"),
    ("ml_risk",          "Risk Score",        "J", 10, "int"),
    ("ml_vol_tier",      "Vol Tier",          "J",  9, None),
    ("ml_line_bucket",   "Line Bucket",       "J", 11, None),
    ("ml_momentum_z",    "Momentum Z",        "J", 11, "dec2"),
    ("ml_opp_strength",  "Opp Strength",      "J", 12, None),
    ("ml_form_trend",    "Form Trend",        "J", 11, None),

    # ── K. REASONING ──────────────────────────────────────────────────────────
    ("preMatchReason",   "Pre-Match Reason",  "K", 80, None),
    ("postMatchReason",  "Post-Match Reason", "K", 80, None),
]

# ── Schema definitions (shown on the Schema sheet) ────────────────────────────
SCHEMA_DEFS = {
    "Date":             "Game date (YYYY-MM-DD)",
    "L3":               "Average pts over last 3 games before this date",
    "L5":               "Average pts over last 5 games",
    "L10":              "Average pts over last 10 games",
    "L20":              "Average pts over last 20 games",
    "L30":              "Season baseline — average pts over last 30 games",
    "Momentum":         "L5 − L30: positive = on a hot streak",
    "Reversion":        "L10 − L30: medium-term deviation from baseline",
    "Accel":            "L3 − L5: is the streak accelerating or fading?",
    "Volume":           "L30 − Line: how far above/below the line is his baseline?",
    "DVP Rank":         "Defence vs Position rank 1-30 (1=toughest, 30=easiest)",
    "Pace Rank":        "Opponent pace rank 1-30 (30=fastest, most possessions)",
    "H2H Avg Pts":      "Average pts scored vs this specific opponent historically",
    "H2H TS Dev":       "True shooting % deviation vs overall in H2H matchups",
    "Elite Prob":       "Elite meta-model probability (0-1). Threshold: ≥0.81=APEX",
    "Pred Error":       "actual_pts − predPts — how far off the projection was",
    "ML Correct":       "1=WIN, 0=LOSS, blank=not yet graded. Primary ML target",
    "Dir Correct":      "1 if prediction direction (OVER/UNDER) was right",
    "Close Call":       "1 if |actual − line| ≤ 2 pts",
    "In Q Band":        "1 if actual_pts fell inside Q25-Q75 prediction band",
    "Consensus Scr":    "Sum of: all_clf_agree + reg_consensus + v12_extreme (0-3)",
    "Risk Score":       "Sum of risk flags: b2b + long_rest + extreme_hot/cold + mean_rev",
    "Vol Tier":         "Volatility tier: LOW(σ≤4) / MED(4-8) / HIGH(σ>8)",
    "Line Bucket":      "Line bracket: 10-14 / 15-19 / 20-24 / 25-29 / 30+",
    "Momentum Z":       "Momentum normalised by std10 — z-score of current streak",
    "Opp Strength":     "DVP-based label: WEAK / MID / TOUGH defence",
    "Form Trend":       "UP if momentum > 2 and accel > 0, DOWN if opposite, else FLAT",
}


# ─────────────────────────────────────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def _safe(v: Any, default=None):
    """Return v, replacing NaN/Inf with default."""
    if v is None:
        return default
    if isinstance(v, float) and (math.isnan(v) or math.isinf(v)):
        return default
    return v


def _pct(v: Any) -> float | None:
    """Convert ratio to percentage (0.45 → 45.0)."""
    v = _safe(v)
    if v is None:
        return None
    return round(float(v) * 100, 1)


def _derive_ml_signals(p: dict) -> dict:
    """
    Compute derived ML training signals from a play dict.
    These are not stored in the JSON but are valuable for model training.
    """
    result    = p.get("result", "")
    actual    = _safe(p.get("actualPts"))
    pred      = _safe(p.get("predPts"))
    line      = float(p.get("line", 0) or 0)
    direction = p.get("direction", p.get("dir", "OVER"))
    q25       = _safe(p.get("q25_v12"))
    q75       = _safe(p.get("q75_v12"))
    momentum  = float(p.get("momentum", 0) or 0)
    std10     = float(p.get("std10", 5) or 5)
    accel     = float(p.get("acceleration", p.get("l3", line) - p.get("l5", line)) or 0)
    dvp       = int(p.get("defP_dynamic", p.get("defP", 15)) or 15)
    is_b2b    = bool(p.get("is_b2b", False))
    long_rest = bool(p.get("is_long_rest", False))
    ext_hot   = bool(p.get("extreme_hot", False))
    ext_cold  = bool(p.get("extreme_cold", False))
    mean_rev  = float(p.get("mean_reversion_risk", 0) or 0)
    all_ag    = bool(p.get("all_clf_agree", False))
    reg_con   = bool(p.get("reg_consensus", False))
    v12_ext   = bool(p.get("v12_extreme", False))
    flags     = int(p.get("flags", 0) or 0)

    # Primary ML target
    ml_correct = None
    if result == "WIN":
        ml_correct = 1
    elif result == "LOSS":
        ml_correct = 0

    # Prediction quality
    pred_error   = round(actual - pred, 1) if actual is not None and pred is not None else None
    dir_correct  = None
    if actual is not None:
        pred_over = (pred or line) > line
        act_over  = actual > line
        dir_correct = 1 if pred_over == act_over else 0

    close_call   = None
    within_q     = None
    if actual is not None:
        close_call = 1 if abs(actual - line) <= 2 else 0
        if q25 is not None and q75 is not None:
            within_q = 1 if q25 <= actual <= q75 else 0

    # Signal consensus score (0-3) — higher = stronger agreement across systems
    consensus = int(all_ag) + int(reg_con) + int(v12_ext)

    # Risk score (0-5) — how many risk factors were present
    risk = int(is_b2b) + int(long_rest) + int(ext_hot) + int(ext_cold) + int(mean_rev > 0.8)

    # Volatility tier
    if std10 <= 4:
        vol_tier = "LOW"
    elif std10 <= 8:
        vol_tier = "MED"
    else:
        vol_tier = "HIGH"

    # Line bucket (useful for stratified analysis)
    if line < 15:
        line_bucket = "10-14"
    elif line < 20:
        line_bucket = "15-19"
    elif line < 25:
        line_bucket = "20-24"
    elif line < 30:
        line_bucket = "25-29"
    else:
        line_bucket = "30+"

    # Momentum z-score
    momentum_z = round(momentum / max(std10, 0.5), 2)

    # Opponent strength label
    if dvp <= 8:
        opp_strength = "TOUGH"
    elif dvp <= 18:
        opp_strength = "MID"
    else:
        opp_strength = "WEAK"

    # Form trend
    if momentum > 2 and accel > 0:
        form_trend = "UP"
    elif momentum < -2 and accel < 0:
        form_trend = "DOWN"
    else:
        form_trend = "FLAT"

    # Actual FG% if box stats present
    actual_fga = _safe(p.get("actual_fga"))
    actual_fgm = _safe(p.get("actual_fgm"))
    actual_fg_pct = None
    if actual_fga and actual_fgm is not None and actual_fga > 0:
        actual_fg_pct = round(actual_fgm / actual_fga * 100, 1)

    return {
        "ml_correct":    ml_correct,
        "ml_pred_error": pred_error,
        "ml_dir_correct":dir_correct,
        "ml_close_call": close_call,
        "ml_within_q":   within_q,
        "ml_consensus":  consensus,
        "ml_risk":       risk,
        "ml_vol_tier":   vol_tier,
        "ml_line_bucket":line_bucket,
        "ml_momentum_z": momentum_z,
        "ml_opp_strength": opp_strength,
        "ml_form_trend": form_trend,
        "actual_fg_pct": actual_fg_pct,
        # Derived reversion & accel (sometimes missing from play dict)
        "reversion":     round(float(p.get("l10", line) or line) - float(p.get("l30", line) or line), 1),
        "acceleration":  round(float(p.get("l3", line) or line) - float(p.get("l5", line) or line), 1),
        "season_progress": _safe(p.get("season_progress", p.get("seasonProgress"))),
        "h2h_trend":     _safe(p.get("h2h_trend")),
    }


def _play_to_row(p: dict) -> dict:
    """Convert a play dict to a flat Excel row dict."""
    ml = _derive_ml_signals(p)

    # ── Normalise camelCase → snake_case before merging ──────────────────
    # play dicts use camelCase in some places; COLUMNS uses snake_case
    normalised = dict(p)  # copy
    _aliases = {
        # (canonical_snake_key, [camelCase fallbacks])
        "over_odds":        ["overOdds"],
        "under_odds":       ["underOdds"],
        "is_home":          ["isHome"],
        "season_progress":  ["seasonProgress"],
        "level_ewm":        ["level_ewm"],   # same but ensure it's there
        "extreme_hot":      ["extreme_hot"],
        "extreme_cold":     ["extreme_cold"],
        "home_away_split":  ["home_away_split"],
        "usage_l30":        ["usage_l30"],
        "min_cv":           ["min_cv"],
        "actual_min":       ["actual_min"],
        "actual_fga":       ["actual_fga"],
        "actual_fgm":       ["actual_fgm"],
    }
    for canonical, fallbacks in _aliases.items():
        if normalised.get(canonical) is None:
            for fb in fallbacks:
                v = normalised.get(fb)
                if v is not None:
                    normalised[canonical] = v
                    break

    # ── Compute fields that are never set explicitly ───────────────────────
    # line_spread = max_line - min_line
    mn = normalised.get("min_line"); mx = normalised.get("max_line")
    if normalised.get("line_spread") is None and mn is not None and mx is not None:
        try:
            normalised["line_spread"] = round(float(mx) - float(mn), 2)
        except Exception:
            pass

    # team — use ptm (player's team abbreviation from game log)
    if normalised.get("team") is None:
        normalised["team"] = normalised.get("ptm") or normalised.get("playerTeam") or None

    # h_a_split = homeAvgPts - awayAvgPts
    if normalised.get("home_away_split") is None:
        h = _safe(normalised.get("homeAvgPts")); a = _safe(normalised.get("awayAvgPts"))
        if h is not None and a is not None:
            try:
                normalised["home_away_split"] = round(float(h) - float(a), 1)
            except Exception:
                pass

    # Merge ML signals into the play for lookup convenience
    merged = {**normalised, **ml}

    row = {}
    for field, label, group, width, fmt in COLUMNS:
        v = _safe(merged.get(field))

        # Type coercions
        if fmt == "pct" and v is not None:
            # Store as decimal (Excel will format as %)
            try:
                v = round(float(v), 4)
                # If value > 1 it's already a percentage (0-100) — normalise
                if v > 1.5:
                    v = round(v / 100, 4)
            except (TypeError, ValueError):
                v = None
        elif fmt == "int" and v is not None:
            try:
                v = int(v)
            except (TypeError, ValueError):
                v = None
        elif fmt in ("dec1", "dec2", "dec3", "dec4") and v is not None:
            dp = int(fmt[-1])
            try:
                v = round(float(v), dp)
            except (TypeError, ValueError):
                v = None
        elif isinstance(v, bool):
            v = int(v)  # 1/0 for Excel booleans (easier to filter/sum)
        elif isinstance(v, (list, dict)):
            v = None  # skip complex types (recent20 etc.)

        row[label] = v

    return row


# ─────────────────────────────────────────────────────────────────────────────
# EXCEL WRITE HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def _col_letter(idx: int) -> str:
    """0-based column index → Excel column letter(s)."""
    result = ""
    idx += 1
    while idx:
        idx, r = divmod(idx - 1, 26)
        result = chr(65 + r) + result
    return result


def _apply_header_formatting(ws, headers: list[str]) -> None:
    """Apply colour-coded group headers, bold, freeze, autofilter."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    # Build label → group map
    label_group = {label: group for _, label, group, _, _ in COLUMNS}

    thin = Border(
        bottom=Side(style="thin"),
        right=Side(style="thin"),
    )

    for ci, label in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=ci)
        group = label_group.get(label, "K")
        bg = _C.get(group, "FFFFFF")
        cell.value = label
        cell.font = Font(bold=True, size=9, name="Calibri")
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin

    ws.row_dimensions[1].height = 32
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{_col_letter(len(headers)-1)}1"


def _apply_column_widths(ws) -> None:
    """Set column widths from schema."""
    label_width = {label: width for _, label, group, width, _ in COLUMNS}
    for ci, col in enumerate(ws.iter_cols(min_row=1, max_row=1), start=1):
        label = col[0].value
        width = label_width.get(label, 10)
        ws.column_dimensions[_col_letter(ci - 1)].width = width


def _apply_result_row_colours(ws, result_col_idx: int, data_start_row: int = 2) -> None:
    """
    Colour WIN rows green, LOSS rows red in the Result column only.
    Kept light so the other columns remain readable.
    """
    from openpyxl.styles import PatternFill
    win_fill  = PatternFill("solid", fgColor=_C["I_win"])
    loss_fill = PatternFill("solid", fgColor=_C["I_loss"])

    for row in ws.iter_rows(min_row=data_start_row, max_row=ws.max_row,
                             min_col=result_col_idx, max_col=result_col_idx):
        cell = row[0]
        if cell.value == "WIN":
            cell.fill = win_fill
        elif cell.value == "LOSS":
            cell.fill = loss_fill


# ─────────────────────────────────────────────────────────────────────────────
# SCHEMA SHEET
# ─────────────────────────────────────────────────────────────────────────────

def _build_schema_sheet(ws) -> None:
    from openpyxl.styles import Font, PatternFill, Alignment

    ws.title = "Schema"
    headers = ["Column", "Group", "Description", "Format"]
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci)
        cell.value = h
        cell.font = Font(bold=True, size=9)
        cell.fill = PatternFill("solid", fgColor="4472C4")
        cell.font = Font(bold=True, size=9, color="FFFFFF")

    for ri, (field, label, group, width, fmt) in enumerate(COLUMNS, start=2):
        ws.cell(row=ri, column=1).value = label
        ws.cell(row=ri, column=2).value = group
        ws.cell(row=ri, column=3).value = SCHEMA_DEFS.get(label, "")
        ws.cell(row=ri, column=4).value = fmt or "general"
        bg = _C.get(group, "FFFFFF")
        for ci in range(1, 3):
            ws.cell(row=ri, column=ci).fill = PatternFill("solid", fgColor=bg)

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 8
    ws.column_dimensions["C"].width = 65
    ws.column_dimensions["D"].width = 10


# ─────────────────────────────────────────────────────────────────────────────
# PUBLIC API
# ─────────────────────────────────────────────────────────────────────────────

def plays_to_df(plays: list[dict]) -> pd.DataFrame:
    """Convert a list of play dicts to a pandas DataFrame ready for Excel."""
    rows = [_play_to_row(p) for p in plays]
    headers = [label for _, label, _, _, _ in COLUMNS]
    df = pd.DataFrame(rows, columns=headers)
    return df


def _dedup_plays_df(df: pd.DataFrame) -> pd.DataFrame:
    """
    Final deduplication guard on the combined DataFrame.
    Key: (Player, Date, Direction) — one row per player per game per bet direction.

    Priority (which row to keep when duplicates exist):
      1. Graded rows (WIN/LOSS/DNP/PUSH) beat ungraded
      2. Among graded, keep the one with highest Elite Prob (most confident prediction)
      3. Among equal confidence, keep the last row (most recently appended)

    This handles:
      - generate() called twice (full dataset written twice)
      - Same player regraded on a different run date
      - Line change creating two predictions for same player/date
    """
    if df.empty:
        return df

    GRADED = {"WIN", "LOSS", "DNP", "PUSH"}
    result_col = "Result" if "Result" in df.columns else None
    prob_col   = "Elite Prob" if "Elite Prob" in df.columns else None
    player_col = "Player" if "Player" in df.columns else None
    date_col   = "Date" if "Date" in df.columns else None
    dir_col    = "Direction" if "Direction" in df.columns else None

    if not all([player_col, date_col]):
        return df  # can't dedup without key columns

    # Build composite dedup score: graded=1 else 0, then elite_prob
    def _grade_score(r):
        result = str(r.get(result_col, "") or "") if result_col else ""
        prob   = float(r.get(prob_col, 0) or 0) if prob_col else 0
        graded = 1 if result in GRADED else 0
        return (graded, prob)

    key_cols = [c for c in [player_col, date_col, dir_col] if c]
    df = df.copy()
    df["_dedup_key"] = df[key_cols].astype(str).agg("|".join, axis=1)

    if result_col:
        df["_grade_score"] = df[result_col].apply(
            lambda x: 1 if str(x) in GRADED else 0
        )
    else:
        df["_grade_score"] = 0

    if prob_col:
        df["_prob_score"] = pd.to_numeric(df[prob_col], errors="coerce").fillna(0)
    else:
        df["_prob_score"] = 0

    # Sort so the best row (graded, highest prob) comes last → keep='last'
    df = df.sort_values(
        ["_dedup_key", "_grade_score", "_prob_score"],
        ascending=[True, True, True]
    ).drop_duplicates(subset=["_dedup_key"], keep="last")

    df = df.drop(columns=["_dedup_key", "_grade_score", "_prob_score"])
    return df.reset_index(drop=True)


def write_ml_dataset(plays: list[dict], verbose: bool = True) -> None:
    """
    Full rebuild — called by generate_season_json.
    Overwrites the existing Excel file with all plays sorted chronologically.
    """
    import openpyxl

    FILE_ML.parent.mkdir(parents=True, exist_ok=True)

    if verbose:
        print(f"\n  [ML Dataset] Building from {len(plays):,} plays...")

    # Sort chronologically
    plays_sorted = sorted(plays, key=lambda p: (p.get("date", ""), p.get("player", "")))
    df = plays_to_df(plays_sorted)

    # Final dedup guard — handles generate() being called twice
    before = len(df)
    df = _dedup_plays_df(df)
    if len(df) < before and verbose:
        print(f"  [ML Dataset] Dedup removed {before - len(df)} duplicate rows")

    graded    = df[df["Result"].isin(["WIN", "LOSS"])]
    wins      = (graded["Result"] == "WIN").sum()
    hr        = wins / len(graded) * 100 if len(graded) > 0 else 0

    # Write via openpyxl for full formatting control
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Plays"

    headers = list(df.columns)
    _apply_header_formatting(ws, headers)

    for row in df.itertuples(index=False, name=None):
        ws.append(list(row))

    _apply_column_widths(ws)

    # Colour Result column
    result_idx = headers.index("Result") + 1
    _apply_result_row_colours(ws, result_idx)

    # Schema sheet
    schema_ws = wb.create_sheet("Schema")
    _build_schema_sheet(schema_ws)

    wb.save(FILE_ML)

    if verbose:
        print(f"  ✓ ML Dataset saved: {FILE_ML.name}")
        print(f"    {len(df):,} plays | {len(graded):,} graded | HR: {hr:.1f}%")
        print(f"    {df.columns.size} columns | {df['Season'].nunique() if 'Season' in df else '?'} seasons")


def append_ml_dataset(plays: list[dict], date_str: str, verbose: bool = True) -> None:
    """
    Daily append — called by batch0_grade after grading.
    Loads existing file, replaces any rows for date_str with the new graded data,
    then saves back. Deduplication key: (Player, Date, Line).
    """
    import openpyxl

    FILE_ML.parent.mkdir(parents=True, exist_ok=True)

    if not FILE_ML.exists():
        if verbose:
            print(f"  [ML Dataset] No existing file — creating fresh.")
        write_ml_dataset(plays, verbose=verbose)
        return

    if verbose:
        print(f"\n  [ML Dataset] Appending {len(plays)} plays for {date_str}...")

    # Load existing data
    try:
        existing_df = pd.read_excel(FILE_ML, sheet_name="Plays")
    except Exception as e:
        if verbose:
            print(f"  ⚠ ML Dataset load failed ({e}) — rebuilding.")
        write_ml_dataset(plays, verbose=verbose)
        return

    # Remove any existing rows for this date (replace with fresh graded data)
    if "Date" in existing_df.columns:
        existing_df = existing_df[existing_df["Date"].astype(str) != date_str]

    # Build new rows for today
    new_df = plays_to_df(plays)

    # Combine and sort
    combined = pd.concat([existing_df, new_df], ignore_index=True)
    combined = combined.sort_values(
        by=["Date", "Player"] if "Player" in combined.columns else ["Date"],
        ascending=True
    ).reset_index(drop=True)

    # Final dedup guard — catches cross-date duplicates, double-runs, line changes
    before = len(combined)
    combined = _dedup_plays_df(combined)
    if len(combined) < before and verbose:
        print(f"  [ML Dataset] Dedup removed {before - len(combined)} duplicate rows")

    # Rewrite with full formatting
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Plays"

    headers = list(combined.columns)
    _apply_header_formatting(ws, headers)

    for row in combined.itertuples(index=False, name=None):
        ws.append(list(row))

    _apply_column_widths(ws)

    if "Result" in headers:
        result_idx = headers.index("Result") + 1
        _apply_result_row_colours(ws, result_idx)

    schema_ws = wb.create_sheet("Schema")
    _build_schema_sheet(schema_ws)

    wb.save(FILE_ML)

    graded = combined[combined["Result"].isin(["WIN", "LOSS"])]
    wins   = (graded["Result"] == "WIN").sum()
    hr     = wins / len(graded) * 100 if len(graded) > 0 else 0

    if verbose:
        print(f"  ✓ ML Dataset updated: {FILE_ML.name}")
        print(f"    {len(combined):,} total plays | {len(graded):,} graded | HR: {hr:.1f}%")


if __name__ == "__main__":
    print("ml_dataset.py — import and call write_ml_dataset(plays) or append_ml_dataset(plays, date)")
